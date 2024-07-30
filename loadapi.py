import torch
from torchvision.models import efficientnet_b0,resnext50_32x4d,mobilenet_v2,efficientnet_b1
from PIL import Image
from torchvision import transforms
import numpy as np
import matplotlib.pyplot as plt
import numpy as np
import io,os
from stl import mesh
import statistics 
import openpyxl
from openpyxl import load_workbook

device = torch.device('cpu')

def xyplane(filename,alpha,beta):
    your_mesh = mesh.Mesh.from_file(filename)

    Zmin = np.min(your_mesh.z)
    Zmax = np.max(your_mesh.z)
    print(f"Minimum Z value in the mesh is {Zmin}, and maximum Z value is {Zmax}")

    num_planes = 30  # Number of XY planes to plot
    #alpha = 0.4
    #beta = 0.7
    z_planes = np.linspace(Zmin + (Zmax - Zmin) * alpha, Zmin + (Zmax - Zmin) * beta, num_planes)

    all_points = []

    for z_plane in z_planes:
        # Get the vertices of the mesh
        vertices = your_mesh.vectors.reshape(-1, 3)

        # Filter vertices based on the Z-plane
        mask = (vertices[:, 2] >= z_plane - 0.01) & (vertices[:, 2] <= z_plane + 0.01)
        selected_vertices = vertices[mask]

        # Add selected vertices to all_points
        all_points.extend(selected_vertices[:, :2])

    all_points = np.array(all_points)

    if len(all_points) > 0:
        fig, ax = plt.subplots(figsize=(8, 8))
        ax.scatter(all_points[:, 0], all_points[:, 1], s=1, color='black', alpha=0.5)
        ax.set_aspect('equal')
        ax.axis('off')  # Remove the axes
        plt.tight_layout()

    # Save the plot to a bytes buffer
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', pad_inches=0, dpi=300)
    plt.close(fig)
    buf.seek(0)

    # Create a PIL Image from the bytes buffer
    image = Image.open(buf)
    image = image.convert('RGB')

    print("XY plane plot generated")
    return image


def xyplane_fast(filename, alpha, beta):
    your_mesh = mesh.Mesh.from_file(filename)

    Zmin = np.min(your_mesh.z)
    Zmax = np.max(your_mesh.z)
    print(f"Minimum Z value in the mesh is {Zmin}, and maximum Z value is {Zmax}")

    num_planes = 20  # Number of XY planes to plot
    z_planes = np.linspace(Zmin + (Zmax - Zmin) * alpha, Zmin + (Zmax - Zmin) * beta, num_planes)

    all_points = []

    # Vectorized operations
    triangles = your_mesh.vectors
    z_coords = triangles[:, :, 2]

    for z_plane in z_planes:
        mask = (np.min(z_coords, axis=1) <= z_plane) & (np.max(z_coords, axis=1) >= z_plane)
        valid_triangles = triangles[mask]

        for triangle in valid_triangles:
            for i in range(3):
                j = (i + 1) % 3
                if (triangle[i, 2] <= z_plane <= triangle[j, 2]) or (triangle[j, 2] <= z_plane <= triangle[i, 2]):
                    t = (z_plane - triangle[i, 2]) / (triangle[j, 2] - triangle[i, 2])
                    point = triangle[i] + t * (triangle[j] - triangle[i])
                    all_points.append(point[:2])

    all_points = np.array(all_points)

    if len(all_points) > 0:
        fig, ax = plt.subplots(figsize=(8, 8))
        ax.scatter(all_points[:, 0], all_points[:, 1], s=5, color='black')
        ax.set_aspect('equal')
        ax.axis('off')  # Remove the axes
        plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', pad_inches=0, dpi=300)
    plt.close(fig)
    buf.seek(0)

    # Create a PIL Image from the bytes buffer
    image = Image.open(buf)
    image = image.convert('RGB')

    print("XY plane plot generated")
    return image

# Initialize the EfficientNet model architecture
def model1(input_image):
    model = efficientnet_b0()

    # Modify the classifier layer to match the number of classes in your task
    num_classes = 16  # Replace with the number of classes in your classification task
    model.classifier[1] = torch.nn.Linear(model.classifier[1].in_features, num_classes)

    # Load the saved model state dictionary
    state_dict = torch.load('720_efficientv0_size512_epoch25_0.0001.pth',  map_location=torch.device('cpu'))

    # Load the state dictionary into the model
    model.load_state_dict(state_dict)

    # Set the model to evaluation mode
    model.eval()

    # Define the data transformations for inference
    inference_transforms = transforms.Compose([
        transforms.Resize((512, 512)),
        transforms.ToTensor(),
    ])

    # Ask the user for the path to the STL file
    #stl_file_path = input("Enter the path to the STL file: ")
    
    # Generate the XY plane plot and get the image
    #input_image = xyplane(stl_file_path)
    
    if input_image:
        # Preprocess the input image
        input_tensor = inference_transforms(input_image)
        input_batch = input_tensor.unsqueeze(0)

        # Move the input batch to the same device as the model
        input_batch = input_batch.to(device)
        model.to(device)
    
        # Perform inference
        with torch.no_grad():
            outputs = model(input_batch)
    
        # Get the predicted class probabilities
        probabilities = torch.softmax(outputs, dim=1)

        # Get the predicted class index
        _, predicted_index = torch.max(probabilities, 1)
        #print(predicted_index)

        # Map the predicted index to the corresponding class label
        class_labels = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '16']  # Replace with your class labels
        predicted_label = class_labels[predicted_index.item()]
    
        #print('Model1 Predicted label:', predicted_label)
    else:
        print("Failed to generate the XY plane plot.")
    return predicted_label

# Initialize the resNext model architecture
def model_4567(input_image):
    model = efficientnet_b0()
    # Modify the classifier layer to match the number of classes in your task
    num_classes = 16  # Replace with the number of classes in your classification task
    model.classifier[1] = torch.nn.Linear(model.classifier[1].in_features, num_classes)
    
   # Load the saved model state dictionary
    state_dict = torch.load('4567_efficientv0_size512_epoch25_0.0001.pth')

    # Load the state dictionary into the model
    model.load_state_dict(state_dict)

    # Set the model to evaluation mode
    model.eval()

    # Define the data transformations for inference
    inference_transforms = transforms.Compose([
        transforms.Resize((512, 512)),
        transforms.ToTensor(),
    ])

    # Ask the user for the path to the STL file
    #stl_file_path = input("Enter the path to the STL file: ")
    
    # Generate the XY plane plot and get the image
    #input_image = xyplane(stl_file_path)
    
    if input_image:
        # Preprocess the input image
        input_tensor = inference_transforms(input_image)
        input_batch = input_tensor.unsqueeze(0)

        # Move the input batch to the same device as the model
        input_batch = input_batch.to(device)
        model.to(device)
    
        # Perform inference
        with torch.no_grad():
            outputs = model(input_batch)
    
        # Get the predicted class probabilities
        probabilities = torch.softmax(outputs, dim=1)

        # Get the predicted class index
        _, predicted_index = torch.max(probabilities, 1)
        #print(predicted_index)

        # Map the predicted index to the corresponding class label
        class_labels = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '16']  # Replace with your class labels
        predicted_label = class_labels[predicted_index.item()]
    
        #print('Model2 Predicted label:', predicted_label)
    else:
        print("Failed to generate the XY plane plot.")
    return predicted_label

# Initialize the mobilenet model architecture
def model_914(input_image):
    model = efficientnet_b0()

    # Modify the classifier layer to match the number of classes in your task
    num_classes = 16  # Replace with the number of classes in your classification task
    model.classifier[1] = torch.nn.Linear(model.classifier[1].in_features, num_classes)
 
    # Load the saved model state dictionary
    state_dict = torch.load('914_efficientnetV0_size512_epoch25_0.0001.pth')

    # Load the state dictionary into the model
    model.load_state_dict(state_dict)

    # Set the model to evaluation mode
    model.eval()

    # Define the data transformations for inference
    inference_transforms = transforms.Compose([
        transforms.Resize((512, 512)),
        transforms.ToTensor(),
    ])

    # Ask the user for the path to the STL file
    #stl_file_path = input("Enter the path to the STL file: ")
    
    # Generate the XY plane plot and get the image
    #input_image = xyplane(stl_file_path)
    
    if input_image:
        # Preprocess the input image
        input_tensor = inference_transforms(input_image)
        input_batch = input_tensor.unsqueeze(0)

        # Move the input batch to the same device as the model
        input_batch = input_batch.to(device)
        model.to(device)
    
        # Perform inference
        with torch.no_grad():
            outputs = model(input_batch)
    
        # Get the predicted class probabilities
        probabilities = torch.softmax(outputs, dim=1)

        # Get the predicted class index
        _, predicted_index = torch.max(probabilities, 1)
        #print(predicted_index)

        # Map the predicted index to the corresponding class label
        class_labels = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '16']  # Replace with your class labels
        predicted_label = class_labels[predicted_index.item()]
    
        #print('Model3 Predicted label:', predicted_label)
    else:
        print("Failed to generate the XY plane plot.")
    return predicted_label

wb = load_workbook('stl.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
    file_number = str(row[0].value)
    if file_number:
        # Construct the filename
        filename = f"{file_number}.stl"
        stl_file_path = os.path.join("new", filename)
        
        print(f"Processing: {stl_file_path}")
        
        if os.path.isfile(stl_file_path):
            input_image1 = xyplane_fast(stl_file_path, 0.1, 0.9)
            #input_image2 = xyplane_fast(stl_file_path, 0, 1)
            if input_image1 is not None:
                label1 = model1(input_image1)
                if label1 >= 4 and label2 <= 7:
                    label1= model_4567(input_image1)
                if label1 >=9:
                    label1 = model_914(input_image1)
                # Update the Excel file with the prediction
                row[1].value = label1
        else:
            print(f"File not found: {stl_file_path}")

# Save the updated Excel file
wb.save('stl_updated.xlsx')
print("Excel file updated with predictions.")

#stl_file_path=input("Enter the path to the STL file: ")


#input_image1= xyplane_fast(stl_file_path,0.1,0.9)
#input_image2= xyplane_fast(stl_file_path,0, 1)
#label1=model1(input_image1)
#label2=model_4567(input_image1)
#label914= model_914(input_image1)
#label3=model1(input_image2)
#input_image2= xyplane(stl_file_path,0.3,0.7)
#label2=model2(input_image2)



#print('Predicted label1:',label1)
#print('Predicted label2:',label2)
#print('Predicted label914:',label914)
#print('Predicted label3:',label3)
#print("final label", statistics.mode([label1,label2,label3,label4,label5,label6]))